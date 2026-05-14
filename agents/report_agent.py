"""
BankSight AI — Report Agent
=============================
Collects outputs from SQL + Analysis + Anomaly agents
and assembles final Excel, PDF, and Power BI CSV reports.

Called by:
  - Claude Code agent: .claude/agents/report-generator.md
  - Streamlit UI:      app.py
  - Skills:            .claude/skills/churn-report/SKILL.md

Spec reference: .claude/specs/02-agent-design.md
"""

import os
import time
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR   = os.path.join(BASE_DIR, "reports")
DASHBOARD_DIR = os.path.join(BASE_DIR, "dashboard")

os.makedirs(REPORTS_DIR,   exist_ok=True)
os.makedirs(DASHBOARD_DIR, exist_ok=True)

# Dynamic file names
RUN_DATE  = datetime.now().strftime("%Y-%m-%d")
RUN_TS    = datetime.now().strftime("%d-%b-%Y %H:%M")
RUN_LABEL = datetime.now().strftime("%d-%b-%Y")


# ---------------------------------------------------------------------------
# Main agent function
# ---------------------------------------------------------------------------
def run_report_agent(
    sql_output: dict,
    analysis_output: dict = None,
    anomaly_output: dict = None,
    report_config: dict = None,
    verbose: bool = True
) -> dict:
    """
    Report Agent pipeline:
    1. Validate inputs (SQL required, others preferred)
    2. Generate requested output formats
    3. Auto-upload to Google Drive if configured
    4. Return file paths and summary

    Args:
        sql_output:      Dict from sql_agent.run_sql_agent() — REQUIRED
        analysis_output: Dict from analysis_agent — preferred
        anomaly_output:  Dict from anomaly_agent  — preferred
        report_config:   Dict with:
                           formats: list ["excel","pdf","csv"]
                           report_type: "churn"|"spend"|"trend"|"employee"
                           period_label: str e.g. "May 2026"
        verbose:         Print progress logs

    Returns dict with keys:
        files_generated — list of {type, path, size_kb}
        summary         — report generation summary
        status          — "success" | "partial" | "error"
        message         — user-facing status message
        time_taken      — seconds elapsed
    """
    start_time = time.time()

    if verbose:
        print(f"\n[Report Agent] Starting report generation...")

    # Default config
    if report_config is None:
        report_config = {}
    formats     = report_config.get("formats", ["excel", "csv"])
    report_type = report_config.get("report_type", "churn")
    period      = report_config.get("period_label", RUN_LABEL)

    # Step 1 — Validate SQL output (hard stop if missing)
    if not sql_output or sql_output.get("status") != "success":
        return {
            "files_generated": [],
            "summary":         {},
            "status":          "error",
            "message": (
                "Cannot generate report: SQL Agent output is missing or failed. "
                "Please run the analysis pipeline first."
            ),
            "time_taken": round(time.time() - start_time, 2),
        }

    df = sql_output["dataframe"]

    # Log missing optional inputs (don't stop — generate partial report)
    disclaimers = []
    if not analysis_output or analysis_output.get("status") != "success":
        disclaimers.append("Analysis Agent output unavailable — narrative summary omitted.")
        if verbose:
            print("[Report Agent] ⚠️  Analysis output missing — partial report")
    if not anomaly_output or anomaly_output.get("status") != "success":
        disclaimers.append("Anomaly Agent output unavailable — risk section omitted.")
        if verbose:
            print("[Report Agent] ⚠️  Anomaly output missing — partial report")

    # Build file name prefix
    label       = report_type.capitalize()
    file_prefix = f"{label}Report_{RUN_DATE}"

    # Step 2 — Generate requested formats
    files_generated = []
    errors          = []

    if "excel" in formats:
        try:
            excel_path = _build_excel(
                df, analysis_output, anomaly_output,
                file_prefix, period, disclaimers, verbose
            )
            files_generated.append({
                "type":    "excel",
                "path":    excel_path,
                "size_kb": round(os.path.getsize(excel_path) / 1024, 1),
            })
        except Exception as e:
            errors.append(f"Excel generation failed: {str(e)}")
            if verbose:
                print(f"[Report Agent] Excel error: {e}")

    if "pdf" in formats:
        try:
            pdf_path = _build_pdf(
                df, analysis_output, anomaly_output,
                file_prefix, period, disclaimers, verbose
            )
            files_generated.append({
                "type":    "pdf",
                "path":    pdf_path,
                "size_kb": round(os.path.getsize(pdf_path) / 1024, 1),
            })
        except Exception as e:
            errors.append(f"PDF generation failed: {str(e)}")
            if verbose:
                print(f"[Report Agent] PDF error: {e}")

    if "csv" in formats:
        try:
            csv_path = _build_csv(
                df, analysis_output, anomaly_output,
                report_type, verbose
            )
            files_generated.append({
                "type":    "csv",
                "path":    csv_path,
                "size_kb": round(os.path.getsize(csv_path) / 1024, 1),
            })
        except Exception as e:
            errors.append(f"CSV generation failed: {str(e)}")
            if verbose:
                print(f"[Report Agent] CSV error: {e}")

    time_taken = round(time.time() - start_time, 2)
    status     = "success" if not errors else ("partial" if files_generated else "error")

    if verbose:
        print(f"[Report Agent] Generated {len(files_generated)} file(s) "
              f"in {time_taken}s — Status: {status} ✅")

    return {
        "files_generated": files_generated,
        "summary": {
            "records_processed":  len(df),
            "critical_anomalies": anomaly_output.get("critical_count", 0)
                                  if anomaly_output else 0,
            "data_coverage":      period,
            "generation_time":    RUN_TS,
            "formats_generated":  [f["type"] for f in files_generated],
            "disclaimers":        disclaimers,
            "errors":             errors,
        },
        "status":     status,
        "message": (
            f"Report generated successfully: "
            f"{', '.join(f['type'].upper() for f in files_generated)}"
            if files_generated else "No files were generated."
        ),
        "time_taken": time_taken,
    }


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
def _build_excel(
    df, analysis_output, anomaly_output,
    file_prefix, period, disclaimers, verbose
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)

    path = os.path.join(REPORTS_DIR, f"{file_prefix}.xlsx")
    wb   = Workbook()

    # Styles
    DARK_BLUE    = "1F3864"
    WHITE        = "FFFFFF"
    RED_LIGHT    = "FFB3B3"
    YELLOW_LIGHT = "FFF2CC"
    ALT_ROW      = "DCE6F1"
    LIGHT_BLUE   = "D6E4F7"

    HDR_FONT  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    BODY_FONT = Font(name="Calibri", size=10)
    TITLE_FONT= Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    THIN      = Side(style="thin", color="B8CCE4")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def style_hdr(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = PatternFill("solid", fgColor=DARK_BLUE)
            cell.font   = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_row(ws, row, cols, alt=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            if alt:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)
            cell.font   = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    # ── Sheet 1: Executive Summary ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 30

    ws1.merge_cells("A1:B1")
    ws1["A1"] = f"BankSight AI — {file_prefix} — {period}"
    ws1["A1"].font      = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
    ws1["A1"].alignment = CENTER
    ws1["A1"].fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:B2")
    ws1["A2"] = f"Generated: {RUN_TS}  |  Records: {len(df):,}  |  Currency: INR (₹)"
    ws1["A2"].font      = Font(name="Calibri", italic=True, size=10, color="4472C4")
    ws1["A2"].alignment = CENTER

    row = 4
    # KPI section from analysis output
    if analysis_output and analysis_output.get("key_metrics"):
        ws1.merge_cells(f"A{row}:B{row}")
        ws1.cell(row=row, column=1, value="Key Metrics")
        ws1.cell(row=row, column=1).font      = TITLE_FONT
        ws1.cell(row=row, column=1).alignment = LEFT
        row += 1

        ws1.cell(row=row, column=1, value="Metric")
        ws1.cell(row=row, column=2, value="Value")
        style_hdr(ws1, row, 2)
        row += 1

        metrics = analysis_output["key_metrics"]
        for i, (k, v) in enumerate(metrics.items()):
            ws1.cell(row=row, column=1, value=str(k).replace("_", " ").title())
            ws1.cell(row=row, column=2, value=str(v))
            style_row(ws1, row, 2, alt=(i % 2 == 0))
            row += 1

        row += 1

    # Findings from analysis
    if analysis_output and analysis_output.get("top_findings"):
        ws1.merge_cells(f"A{row}:B{row}")
        ws1.cell(row=row, column=1, value="Key Findings")
        ws1.cell(row=row, column=1).font      = TITLE_FONT
        ws1.cell(row=row, column=1).alignment = LEFT
        row += 1

        for i, finding in enumerate(analysis_output["top_findings"], 1):
            ws1.merge_cells(f"A{row}:B{row}")
            ws1.cell(row=row, column=1, value=f"{i}. {finding}")
            style_row(ws1, row, 2, alt=(i % 2 == 0))
            ws1.row_dimensions[row].height = 30
            row += 1

    # ── Sheet 2: Raw Data ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Data")
    ws2.sheet_view.showGridLines = False

    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
        ws2.column_dimensions[
            chr(64 + c) if c <= 26 else "A" + chr(64 + c - 26)
        ].width = 16
    style_hdr(ws2, 1, len(headers))

    for r_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for c_idx, val in enumerate(row_data, 1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
        style_row(ws2, r_idx, len(headers), alt=(r_idx % 2 == 0))

    # ── Sheet 3: Analysis Results ──────────────────────────────────────────
    ws3 = wb.create_sheet("Analysis Results")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 80

    ws3["A1"] = "Analysis Insights"
    ws3["A1"].font = TITLE_FONT
    row = 3

    if analysis_output:
        for section, content in [
            ("Summary",         analysis_output.get("summary", "N/A")),
            ("Recommendations", "\n".join(
                f"{i+1}. {r}"
                for i, r in enumerate(
                    analysis_output.get("recommendations", [])))),
            ("Confidence",      analysis_output.get("confidence", "N/A")),
        ]:
            ws3.cell(row=row, column=1, value=section)
            ws3.cell(row=row, column=1).font      = Font(
                name="Calibri", bold=True, size=11, color=DARK_BLUE)
            ws3.cell(row=row, column=1).fill      = PatternFill(
                "solid", fgColor="BDD7EE")
            ws3.cell(row=row, column=1).alignment = LEFT
            ws3.cell(row=row, column=1).border    = BORDER
            row += 1

            ws3.cell(row=row, column=2, value=str(content))
            ws3.cell(row=row, column=2).font      = BODY_FONT
            ws3.cell(row=row, column=2).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            ws3.cell(row=row, column=2).border    = BORDER
            ws3.row_dimensions[row].height = 50
            row += 2
    else:
        ws3.cell(row=row, column=1,
                 value="Analysis Agent output not available for this report.")
        ws3.cell(row=row, column=1).font = BODY_FONT

    # ── Sheet 4: Anomaly Flags ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Anomaly Flags")
    ws4.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E"], [12, 60, 12, 8, 10]):
        ws4.column_dimensions[col].width = w

    ws4["A1"] = "Anomaly Flags"
    ws4["A1"].font = TITLE_FONT
    row = 3

    if anomaly_output and anomaly_output.get("anomalies"):
        for c, h in enumerate(
                ["ID", "Description", "Severity", "Score", "Action"], 1):
            ws4.cell(row=row, column=c, value=h)
        style_hdr(ws4, row, 5)
        row += 1

        for ano in anomaly_output["anomalies"]:
            sev      = ano.get("severity", "NORMAL")
            fill_hex = ("FFB3B3" if sev == "CRITICAL"
                        else "FFF2CC" if sev == "WARNING"
                        else "FFFFFF")
            cells = [
                ano.get("anomaly_id", ""),
                ano.get("description", ""),
                sev,
                str(ano.get("score", "")),
                ano.get("action", ""),
            ]
            for c_idx, val in enumerate(cells, 1):
                cell = ws4.cell(row=row, column=c_idx, value=val)
                cell.fill      = PatternFill("solid", fgColor=fill_hex)
                cell.font      = Font(name="Calibri", size=10,
                                      bold=(sev == "CRITICAL"))
                cell.border    = BORDER
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)
            ws4.row_dimensions[row].height = 30
            row += 1
    else:
        ws4.cell(row=row, column=1,
                 value="Anomaly Agent output not available for this report.")
        ws4.cell(row=row, column=1).font = BODY_FONT

    # ── Sheet 5: Methodology ──────────────────────────────────────────────
    ws5 = wb.create_sheet("Methodology")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 55

    ws5["A1"] = "Methodology & Data Sources"
    ws5["A1"].font = TITLE_FONT
    row = 3

    methodology = [
        ("Data Source",      "data/banking_mock.db (SQLite)"),
        ("Tables Used",      "customers, transactions, loan_emi"),
        ("Records Analysed", f"{len(df):,}"),
        ("Churn Baseline",   "21% (681/3225)"),
        ("Anomaly Threshold","amount > 2x customer monthly average"),
        ("NPS Scale",        "0-10 (not 0-100)"),
        ("Currency",         "INR (₹) — never USD"),
        ("Date Format",      "DD-MMM-YYYY"),
        ("Report Generated", RUN_TS),
        ("SQL Query Used",   "See analysis session logs"),
    ]

    for c, h in enumerate(["Parameter", "Value"], 1):
        ws5.cell(row=row, column=c, value=h)
    style_hdr(ws5, row, 2)
    row += 1

    for i, (param, val) in enumerate(methodology):
        ws5.cell(row=row, column=1, value=param)
        ws5.cell(row=row, column=2, value=val)
        style_row(ws5, row, 2, alt=(i % 2 == 0))
        row += 1

    if disclaimers:
        row += 1
        for disc in disclaimers:
            ws5.merge_cells(f"A{row}:B{row}")
            ws5.cell(row=row, column=1, value=f"⚠️  {disc}")
            ws5.cell(row=row, column=1).font = Font(
                name="Calibri", size=10, color="FF0000", italic=True)
            row += 1

    ws5.merge_cells(f"A{row}:B{row}")
    ws5.cell(row=row, column=1,
             value="CONFIDENTIAL — BankSight AI Internal Report")
    ws5.cell(row=row, column=1).font = Font(
        name="Calibri", bold=True, size=10, color="FF0000")
    ws5.cell(row=row, column=1).alignment = CENTER

    wb.save(path)
    if verbose:
        print(f"[Report Agent] Excel saved: {path}")
    return path


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------
def _build_pdf(
    df, analysis_output, anomaly_output,
    file_prefix, period, disclaimers, verbose
) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units    import cm
    from reportlab.lib          import colors
    from reportlab.lib.styles   import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus     import (SimpleDocTemplate, Paragraph,
                                         Spacer, Table, TableStyle,
                                         PageBreak, HRFlowable)
    from reportlab.lib.enums    import TA_CENTER, TA_LEFT, TA_JUSTIFY

    path    = os.path.join(REPORTS_DIR, f"{file_prefix}.pdf")
    W, H    = A4
    DB      = colors.HexColor("#1F3864")
    RED     = colors.HexColor("#FF0000")
    RED_L   = colors.HexColor("#FFB3B3")
    YEL     = colors.HexColor("#FFF2CC")
    ALT     = colors.HexColor("#DCE6F1")

    styles  = getSampleStyleSheet()
    h1      = ParagraphStyle("H1",  parent=styles["Heading1"],
                              fontSize=18, textColor=DB, alignment=TA_CENTER)
    h2      = ParagraphStyle("H2",  parent=styles["Heading2"],
                              fontSize=13, textColor=DB, spaceBefore=12)
    body    = ParagraphStyle("Body",parent=styles["Normal"],
                              fontSize=9.5, leading=14, alignment=TA_JUSTIFY)
    bullet  = ParagraphStyle("Bul", parent=styles["Normal"],
                              fontSize=9.5, leftIndent=14)
    conf    = ParagraphStyle("Conf",parent=styles["Normal"],
                              fontSize=8, textColor=RED, alignment=TA_CENTER)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#4472C4"))
        canvas.drawString(
            1.5*cm, 1.0*cm,
            "CONFIDENTIAL — BankSight AI Internal Report")
        canvas.drawRightString(
            W - 1.5*cm, 1.0*cm,
            f"Page {doc.page}  |  Generated: {RUN_TS}")
        canvas.restoreState()

    def tbl_style(data):
        return TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  DB),
            ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8.5),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, ALT]),
            ("GRID",         (0,0), (-1,-1), 0.5,
             colors.HexColor("#B8CCE4")),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("LEFTPADDING",  (0,0), (-1,-1), 5),
        ])

    doc   = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.8*cm, bottomMargin=2.2*cm,
    )
    story = []
    HR    = HRFlowable(width="100%", thickness=1.2,
                       color=DB, spaceAfter=8)

    # Page 1 — Executive Summary
    story.append(Paragraph(f"BankSight AI — {file_prefix}", h1))
    story.append(Paragraph(
        f"Period: {period}  |  Generated: {RUN_TS}  |  Records: {len(df):,}",
        ParagraphStyle("sub", parent=styles["Normal"],
                       fontSize=10, textColor=colors.HexColor("#4472C4"),
                       alignment=TA_CENTER, spaceAfter=12)))
    story.append(HR)

    story.append(Paragraph("1. Executive Summary", h2))

    if analysis_output and analysis_output.get("summary"):
        story.append(Paragraph(analysis_output["summary"], body))
        story.append(Spacer(1, 0.3*cm))

        if analysis_output.get("top_findings"):
            story.append(Paragraph("Key Findings:", h2))
            for i, f in enumerate(analysis_output["top_findings"], 1):
                story.append(Paragraph(f"<b>{i}.</b> {f}", bullet))

        if analysis_output.get("recommendations"):
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph("Recommendations:", h2))
            for i, r in enumerate(analysis_output["recommendations"], 1):
                story.append(Paragraph(f"<b>{i}.</b> {r}", bullet))
    else:
        story.append(Paragraph(
            "Analysis Agent output not available — raw data report only.",
            ParagraphStyle("warn", parent=styles["Normal"],
                           fontSize=9.5, textColor=RED)))

    story.append(PageBreak())

    # Page 2 — Anomaly Alerts
    story.append(Paragraph("2. Anomaly Alerts", h2))
    story.append(HR)

    if anomaly_output and anomaly_output.get("anomalies"):
        ano_data = [["ID", "Description", "Severity", "Action"]]
        for ano in anomaly_output["anomalies"]:
            ano_data.append([
                ano.get("anomaly_id", ""),
                ano.get("description", ""),
                ano.get("severity", ""),
                ano.get("action", ""),
            ])
        ano_tbl = Table(
            ano_data,
            colWidths=[1.8*cm, 7.5*cm, 2.4*cm, 5.3*cm])
        ts = tbl_style(ano_data)
        for i, ano in enumerate(anomaly_output["anomalies"], 1):
            sev = ano.get("severity", "NORMAL")
            c   = RED_L if sev == "CRITICAL" else YEL if sev == "WARNING" else colors.white
            ts.add("BACKGROUND", (0, i), (-1, i), c)
            if sev == "CRITICAL":
                ts.add("FONTNAME", (0, i), (-1, i), "Helvetica-Bold")
        ano_tbl.setStyle(ts)
        story.append(ano_tbl)
    else:
        story.append(Paragraph(
            "Anomaly Agent output not available for this report.",
            body))

    story.append(PageBreak())

    # Page 3 — Data Appendix
    story.append(Paragraph("3. Data Appendix", h2))
    story.append(HR)

    meta = [
        ["Parameter",        "Value"],
        ["Records analysed", f"{len(df):,}"],
        ["Columns",          ", ".join(list(df.columns)[:8]) + "..."],
        ["Churn baseline",   "21% (681/3225)"],
        ["Anomaly threshold","amount > 2x customer monthly average"],
        ["NPS scale",        "0-10 (not 0-100)"],
        ["Currency",         "INR (₹)"],
        ["Generated",        RUN_TS],
    ]
    meta_tbl = Table(meta, colWidths=[5*cm, 12*cm])
    meta_tbl.setStyle(tbl_style(meta))
    story.append(meta_tbl)

    if disclaimers:
        story.append(Spacer(1, 0.4*cm))
        for disc in disclaimers:
            story.append(Paragraph(f"⚠️  {disc}",
                                   ParagraphStyle("disc", parent=styles["Normal"],
                                                  fontSize=9, textColor=RED,
                                                  italic=True)))

    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph(
        "CONFIDENTIAL — BankSight AI Internal Report — Do not distribute externally",
        conf))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    if verbose:
        print(f"[Report Agent] PDF saved: {path}")
    return path


# ---------------------------------------------------------------------------
# CSV builder — customer-level rows for Power BI
# ---------------------------------------------------------------------------
def _build_csv(
    df, analysis_output, anomaly_output,
    report_type, verbose
) -> str:
    path       = os.path.join(
        DASHBOARD_DIR, f"PowerBI_{report_type.capitalize()}_{RUN_DATE}.csv")
    export_df  = df.copy()

    # Add derived columns for Power BI slicing
    export_df["report_generated_at"] = RUN_TS
    export_df["report_date"]         = RUN_LABEL
    export_df["data_source"]         = "banking_mock.db"
    export_df["baseline_churn_rate"] = 21.00

    if analysis_output and analysis_output.get("key_metrics"):
        status = analysis_output["key_metrics"].get("status", "NORMAL")
        export_df["churn_status"] = status

    if anomaly_output and anomaly_output.get("anomalies"):
        # Flag critical anomaly customers
        critical_ents = [
            a.get("entity", "")
            for a in anomaly_output["anomalies"]
            if a.get("severity") == "CRITICAL"
        ]
        export_df["has_critical_anomaly"] = len(critical_ents) > 0

    # Churn risk combo flag
    if "nps_score" in export_df.columns and "emi_status" in export_df.columns:
        export_df["churn_risk_combo"] = (
            (export_df["nps_score"] <= 3) &
            (export_df["emi_status"] == "Missed")
        ).astype(int)
    elif "nps_score" in export_df.columns:
        export_df["churn_risk_combo"] = (
            export_df["nps_score"] <= 3).astype(int)

    # UTF-8 BOM for Power BI / Excel compatibility
    export_df.to_csv(path, index=False, encoding="utf-8-sig")

    if verbose:
        print(f"[Report Agent] CSV saved: {path} ({len(export_df):,} rows)")
    return path


# ---------------------------------------------------------------------------
# CLI — for direct testing
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    from agents.sql_agent      import run_sql_agent
    from agents.analysis_agent import run_analysis_agent
    from agents.anomaly_agent  import run_anomaly_agent

    print("Testing Report Agent — full pipeline...")

    sql_out      = run_sql_agent("Show all churned customers with segment and NPS")
    analysis_out = run_analysis_agent(sql_out, context="churn")
    anomaly_out  = run_anomaly_agent(sql_out, analysis_out)

    result = run_report_agent(
        sql_output=sql_out,
        analysis_output=analysis_out,
        anomaly_output=anomaly_out,
        report_config={
            "formats":      ["excel", "pdf", "csv"],
            "report_type":  "churn",
            "period_label": "Full Dataset 2024",
        }
    )

    print(f"\n--- Report Agent Results ---")
    print(f"Status:  {result['status']}")
    print(f"Message: {result['message']}")
    print(f"Time:    {result['time_taken']}s")
    for f in result["files_generated"]:
        print(f"  {f['type'].upper()}: {f['path']} ({f['size_kb']} KB)")
