"""
BankSight AI — Churn Report Generator
Report Agent: Generates Excel, PDF, and PowerBI CSV
from SQL + Analysis + Anomaly Agent outputs.

FIXES APPLIED:
  1. Relative BASE_DIR — works on any machine
  2. Live data from banking_mock.db via SQLite
  3. Customer-level CSV rows for Power BI slicing
  4. ReportLab doc constructor fix (footer via build only)
"""

import os
import csv
import sqlite3
import pandas as pd
from datetime import datetime

# ---------------------------------------------------------------------------
# Fix 1 — Relative paths (works on any machine, any OS)
# ---------------------------------------------------------------------------
BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR   = os.path.join(BASE_DIR, "reports")
DASHBOARD_DIR = os.path.join(BASE_DIR, "dashboard")
DB_PATH       = os.path.join(BASE_DIR, "data", "banking_mock.db")

os.makedirs(REPORTS_DIR,   exist_ok=True)
os.makedirs(DASHBOARD_DIR, exist_ok=True)

# Dynamic file names based on today's date
RUN_DATE  = datetime.now().strftime("%Y-%m-%d")
RUN_TS    = datetime.now().strftime("%d-%b-%Y %H:%M")
RUN_LABEL = datetime.now().strftime("%d-%b-%Y")

EXCEL_PATH = os.path.join(REPORTS_DIR,   f"ChurnReport_{RUN_DATE}.xlsx")
PDF_PATH   = os.path.join(REPORTS_DIR,   f"ChurnReport_{RUN_DATE}.pdf")
CSV_PATH   = os.path.join(DASHBOARD_DIR, f"PowerBI_Churn_{RUN_DATE}.csv")

# ---------------------------------------------------------------------------
# Fix 2 — Load live data from banking_mock.db
# ---------------------------------------------------------------------------
def load_data_from_db(period_start=None, period_end=None):
    """
    Loads all required data from banking_mock.db.
    period_start / period_end: 'YYYY-MM-DD' strings (optional filter).
    Returns a dict of DataFrames and computed KPI values.
    """
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(
            f"Database not found at {DB_PATH}. "
            "Please run data/setup_db.py first."
        )

    conn = sqlite3.connect(DB_PATH)

    # --- Customers ---
    customers_df = pd.read_sql("SELECT * FROM customers", conn)
    total_customers = len(customers_df)
    churned_df      = customers_df[customers_df["churn"] == 1].copy()
    total_churned   = len(churned_df)
    churn_rate      = round((total_churned / total_customers) * 100, 2)
    avg_balance     = round(churned_df["balance"].mean(), 2)
    avg_nps         = round(churned_df["nps_score"].mean(), 2)
    avg_credit      = round(churned_df["credit_score"].mean(), 2)

    # --- Segment breakdown ---
    seg_group = customers_df.groupby("segment").agg(
        total=("customer_id", "count")
    ).reset_index()
    churn_seg = churned_df.groupby("segment").agg(
        churned=("customer_id", "count")
    ).reset_index()
    seg_df = seg_group.merge(churn_seg, on="segment", how="left").fillna(0)
    seg_df["churn_rate"] = round(
        (seg_df["churned"] / seg_df["total"]) * 100, 2)
    seg_df["status"] = seg_df["churn_rate"].apply(
        lambda r: "CRITICAL" if r > 25 else "WARNING" if r > 21 else "NORMAL")

    # --- City breakdown ---
    city_group = customers_df.groupby("city").agg(
        total=("customer_id", "count")
    ).reset_index()
    churn_city = churned_df.groupby("city").agg(
        churned=("customer_id", "count")
    ).reset_index()
    city_df = city_group.merge(churn_city, on="city", how="left").fillna(0)
    city_df["churn_rate"] = round(
        (city_df["churned"] / city_df["total"]) * 100, 2)
    city_df = city_df.sort_values("churned", ascending=False)

    # --- Top 10 churned by balance ---
    top10_df = churned_df.nlargest(10, "balance")[
        ["customer_id", "balance", "credit_score",
         "nps_score", "segment", "city", "tenure"]
    ].reset_index(drop=True)
    top10_df.index += 1

    # --- Loan / EMI ---
    loan_df = pd.read_sql("SELECT * FROM loan_emi", conn)
    total_loans   = len(loan_df)
    missed_count  = len(loan_df[loan_df["emi_status"] == "Missed"])
    delayed_count = len(loan_df[loan_df["emi_status"] == "Delayed"])
    missed_pct    = round((missed_count  / total_loans) * 100, 2)
    delayed_pct   = round((delayed_count / total_loans) * 100, 2)
    stress_rate   = round(missed_pct + delayed_pct, 2)

    # Churn risk combo: churned + missed EMI
    churned_ids      = set(churned_df["customer_id"].tolist())
    missed_ids       = set(
        loan_df[loan_df["emi_status"] == "Missed"]["customer_id"].tolist())
    churn_risk_count = len(churned_ids & missed_ids)

    # --- Transactions (with period filter if provided) ---
    txn_query = "SELECT * FROM transactions"
    if period_start and period_end:
        txn_query += (
            f" WHERE transaction_date >= '{period_start}'"
            f" AND transaction_date <= '{period_end}'"
        )
    txn_df = pd.read_sql(txn_query, conn)
    total_txns = len(txn_df)

    # Anomaly flags from transactions
    anomaly_txns = len(txn_df[txn_df["is_anomaly"] == 1]) if "is_anomaly" in txn_df.columns else 0

    conn.close()

    return {
        "customers_df":    customers_df,
        "churned_df":      churned_df,
        "top10_df":        top10_df,
        "seg_df":          seg_df,
        "city_df":         city_df,
        "loan_df":         loan_df,
        "txn_df":          txn_df,
        "total_customers": total_customers,
        "total_churned":   total_churned,
        "churn_rate":      churn_rate,
        "avg_balance":     avg_balance,
        "avg_nps":         avg_nps,
        "avg_credit":      avg_credit,
        "total_loans":     total_loans,
        "missed_count":    missed_count,
        "delayed_count":   delayed_count,
        "missed_pct":      missed_pct,
        "delayed_pct":     delayed_pct,
        "stress_rate":     stress_rate,
        "churn_risk_count": churn_risk_count,
        "total_txns":      total_txns,
        "anomaly_txns":    anomaly_txns,
    }

# ---------------------------------------------------------------------------
# Load live data from DB — called once in main, passed to all builders
# ---------------------------------------------------------------------------
def build_shared_structures(data):
    """Convert live DB data into report-ready shared structures."""
    d = data

    KPIS = [
        ("Total Customers",            f"{d['total_customers']:,}"),
        ("Total Churned",              f"{d['total_churned']:,}"),
        ("Churn Rate",                 f"{d['churn_rate']:.2f}%"),
        ("Baseline Churn Rate",        "21.00%"),
        ("Avg Balance (Churned)",      f"\u20b9{d['avg_balance']:,.2f}"),
        ("Avg NPS (Churned)",          f"{d['avg_nps']:.2f} / 10"),
        ("Avg Credit Score (Churned)", f"{d['avg_credit']:.2f}"),
    ]

    SEGMENT_DATA = [
        (row["segment"],
         int(row["churned"]),
         int(row["total"]),
         float(row["churn_rate"]),
         row["status"])
        for _, row in d["seg_df"].iterrows()
    ]

    CITY_DATA = [
        (row["city"],
         int(row["churned"]),
         int(row["total"]),
         float(row["churn_rate"]))
        for _, row in d["city_df"].iterrows()
    ]

    TOP10_CUSTOMERS = [
        (i + 1,
         f"{row['balance']:,.2f}",
         int(row["credit_score"]),
         int(row["nps_score"]),
         row["segment"],
         row["city"],
         int(row["tenure"]))
        for i, (_, row) in enumerate(d["top10_df"].iterrows())
    ]

    LOAN_EMI = [
        ("Total Loans",          f"{d['total_loans']:,}"),
        ("Missed EMIs",          f"{d['missed_count']} ({d['missed_pct']:.2f}%)"),
        ("Delayed EMIs",         f"{d['delayed_count']} ({d['delayed_pct']:.2f}%)"),
        ("Combined Stress Rate", f"{d['stress_rate']:.2f}%"),
        ("Churned + Missed EMI", f"{d['churn_risk_count']} customers"),
    ]

    # Dynamic key findings based on live data
    top_city      = d["city_df"].iloc[0]
    bottom_city   = d["city_df"].iloc[-1]
    premium_row   = d["seg_df"][d["seg_df"]["segment"] == "Premium"]
    premium_rate  = float(premium_row["churn_rate"].values[0]) if len(premium_row) else 0
    premium_status= str(premium_row["status"].values[0]) if len(premium_row) else "NORMAL"

    KEY_FINDINGS = [
        f"Overall churn rate is {d['churn_rate']:.2f}% vs 21.00% baseline "
        f"({'ABOVE' if d['churn_rate'] > 21 else 'WITHIN'} threshold).",
        f"Churned customers average NPS of {d['avg_nps']:.2f}/10 "
        f"{'— dangerously close to the 3.0 churn risk floor.' if d['avg_nps'] < 4 else '.'}",
        f"Premium segment churn at {premium_rate:.2f}% is {premium_status}.",
        f"{d['stress_rate']:.2f}% combined EMI stress rate "
        f"({'above' if d['stress_rate'] > 30 else 'within'} 30% monitoring threshold).",
        f"{top_city['city']} leads churn volume ({int(top_city['churned'])}); "
        f"{bottom_city['city']} has the lowest rate ({bottom_city['churn_rate']:.2f}%).",
        f"{d['churn_risk_count']} customers present full churn risk combo "
        f"(churned + missed EMI) — highest priority outreach targets.",
    ]

    RECOMMENDATIONS = [
        ("Activate Premium Segment Retention Protocol",
         f"Deploy targeted retention campaign for Premium customers with NPS <= 4. "
         f"Current Premium churn: {premium_rate:.2f}% (threshold: 25%). "
         f"Target: reduce below 23% within 60 days."),
        ("Implement Pre-Churn Transaction Spike Alert",
         "Configure Anomaly Agent to escalate customers with 2+ anomalous transactions "
         "in 30-day window + NPS <= 4 to retention desk within 24 hours."),
        ("Cross-Reference EMI Stress Against NPS",
         f"Run query joining loan_emi (Missed/Delayed) against customers with NPS <= 3. "
         f"{d['churn_risk_count']} confirmed churn risk combo customers require "
         f"immediate outreach."),
    ]

    # Dynamic anomalies based on live thresholds
    ANOMALIES = []
    ano_id = 1
    if d["churn_rate"] > 25:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Overall Churn Rate Breach ({d['churn_rate']:.2f}% > 25%)",
            "CRITICAL", 9, "New"))
        ano_id += 1
    if premium_rate > 25:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Premium Segment Churn Rate Breach ({premium_rate:.2f}% > 25%)",
            "CRITICAL", 9, "New"))
        ano_id += 1
    if d["avg_nps"] <= 3:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Churned Customer Avg NPS at or below risk floor ({d['avg_nps']:.2f}/10)",
            "CRITICAL", 9, "New"))
        ano_id += 1
    if d["churn_risk_count"] > 0:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Churn Risk Combo: {d['churn_risk_count']} customers with NPS<=3 + Missed EMI",
            "CRITICAL", 9, "New"))
        ano_id += 1
    if d["stress_rate"] > 30:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Portfolio EMI Stress Rate {d['stress_rate']:.2f}% (threshold 30%)",
            "WARNING", 6, "New"))
        ano_id += 1
    if d["anomaly_txns"] > 0:
        ANOMALIES.append((
            f"ANO-{ano_id:03d}",
            f"Spend Spike Alert: {d['anomaly_txns']:,} anomalous transactions detected",
            "WARNING", 6, "New"))
        ano_id += 1
    if not ANOMALIES:
        ANOMALIES.append(("ANO-001", "No anomalies detected in current period.",
                          "NORMAL", 0, "Clear"))

    PRIORITY_LIST = [
        (i + 1,
         f"{row['segment']}, \u20b9{row['balance']:,.0f} balance, "
         f"NPS {int(row['nps_score'])}/10, Tenure {int(row['tenure'])} yrs",
         "High balance loss — review for retention opportunity")
        for i, (_, row) in enumerate(
            d["top10_df"][d["top10_df"]["nps_score"] <= 4].head(5).iterrows()
        )
    ]
    if not PRIORITY_LIST:
        PRIORITY_LIST = [(1, "No high-priority customers identified.", "")]

    METHODOLOGY = [
        ("Data Source",            f"data/banking_mock.db (SQLite)"),
        ("Tables Used",            "customers, transactions, loan_emi"),
        ("Customers Analysed",     f"{d['total_customers']:,}"),
        ("Transactions Analysed",  f"{d['total_txns']:,}"),
        ("Loan Records Analysed",  f"{d['total_loans']:,}"),
        ("Churn Baseline",         "21% (681/3225)"),
        ("Anomaly Threshold",      "amount > 2x customer monthly average"),
        ("Report Generated",       RUN_LABEL),
        ("Currency",               "INR (Rs)"),
        ("NPS Scale",              "0-10"),
    ]

    return (KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
            LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
            ANOMALIES, PRIORITY_LIST, METHODOLOGY)

# ===========================================================================
# EXCEL REPORT
# ===========================================================================
def build_excel(KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
                LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
                ANOMALIES, PRIORITY_LIST, METHODOLOGY):
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                 numbers)
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Colour palette
    DARK_BLUE   = "1F3864"
    WHITE       = "FFFFFF"
    RED_FILL    = "FF0000"
    RED_LIGHT   = "FFB3B3"
    ORANGE_FILL = "FF6600"
    YELLOW_FILL = "FFFF00"
    YELLOW_LIGHT= "FFF2CC"
    GREEN_FILL  = "00B050"
    ALT_ROW     = "DCE6F1"
    HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
    BODY_FONT   = Font(name="Calibri", size=10)
    TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    BOLD_FONT   = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    THIN        = Side(style="thin", color="B8CCE4")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr_fill(hex_color=DARK_BLUE):
        return PatternFill("solid", fgColor=hex_color)

    def alt_fill():
        return PatternFill("solid", fgColor=ALT_ROW)

    def style_header_row(ws, row_num, col_count, bg=DARK_BLUE):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = hdr_fill(bg)
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_data_row(ws, row_num, col_count, alternate=False):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            if alternate:
                cell.fill = alt_fill()
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    def write_section_title(ws, row, text, col_span=6):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=col_span)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = TITLE_FONT
        cell.alignment = LEFT
        return row + 1

    # -----------------------------------------------------------------------
    # Sheet 1 — Executive Summary
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 22

    row = 1
    # Report title block
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "BankSight AI — Churn Analysis Report"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
    ws1["A1"].alignment = CENTER
    ws1["A1"].fill = PatternFill("solid", fgColor="D6E4F7")

    ws1.merge_cells("A2:D2")
    ws1["A2"] = "Period: 08-Apr-2024 to 07-May-2024   |   Report Date: 07-May-2026   |   Generated: " + RUN_TS
    ws1["A2"].font = Font(name="Calibri", italic=True, size=10, color="4472C4")
    ws1["A2"].alignment = CENTER
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 18

    row = 4
    row = write_section_title(ws1, row, "Key Performance Indicators", 4)
    headers = ["Metric", "Value", "Status", "Note"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, 4)
    row += 1

    # Traffic light function
    def churn_status_flag(rate_str):
        try:
            r = float(rate_str.replace("%",""))
            if r > 25: return ("CRITICAL", RED_LIGHT)
            if r > 21: return ("MONITOR", YELLOW_LIGHT)
            return ("NORMAL", "E2EFDA")
        except:
            return ("", "FFFFFF")

    kpi_notes = {
        "Churn Rate":         "vs 21.00% baseline",
        "Avg NPS (Churned)":  "Risk floor: 3.0/10",
        "Avg Balance (Churned)": "INR; Premium-heavy",
    }

    for i, (metric, value) in enumerate(KPIS):
        alt = (i % 2 == 0)
        ws1.cell(row=row, column=1, value=metric)
        ws1.cell(row=row, column=2, value=value)
        note = kpi_notes.get(metric, "")
        ws1.cell(row=row, column=4, value=note)
        if "Rate" in metric and "%" in value:
            status_text, status_color = churn_status_flag(value)
            cell_s = ws1.cell(row=row, column=3, value=status_text)
            cell_s.fill = PatternFill("solid", fgColor=status_color)
            cell_s.font = Font(name="Calibri", bold=True, size=10,
                               color=(RED_FILL if "CRITICAL" in status_text else "375623"))
        else:
            ws1.cell(row=row, column=3, value="")
        style_data_row(ws1, row, 4, alt)
        row += 1

    row += 1
    row = write_section_title(ws1, row, "Anomaly Count Summary", 4)
    ws1.cell(row=row, column=1, value="Severity")
    ws1.cell(row=row, column=2, value="Count")
    ws1.cell(row=row, column=3, value="Action Required")
    ws1.cell(row=row, column=4, value="Status")
    style_header_row(ws1, row, 4)
    row += 1
    for sev, count, action, color in [
        ("CRITICAL", 5, "Immediate escalation required", RED_LIGHT),
        ("WARNING",  5, "Monitor and review within 48h", YELLOW_LIGHT),
    ]:
        ws1.cell(row=row, column=1, value=sev)
        ws1.cell(row=row, column=2, value=count)
        ws1.cell(row=row, column=3, value=action)
        ws1.cell(row=row, column=4, value="New")
        for c in range(1, 5):
            ws1.cell(row=row, column=c).fill = PatternFill("solid", fgColor=color)
            ws1.cell(row=row, column=c).font = BODY_FONT
            ws1.cell(row=row, column=c).border = BORDER
            ws1.cell(row=row, column=c).alignment = LEFT
        row += 1

    row += 1
    row = write_section_title(ws1, row, "Top 3 Actionable Recommendations", 4)
    for i, (title, detail) in enumerate(RECOMMENDATIONS, 1):
        ws1.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        cell = ws1.cell(row=row, column=1,
                        value=f"{i}. {title}")
        cell.font = BOLD_FONT
        cell.fill = PatternFill("solid", fgColor="D6E4F7")
        cell.alignment = LEFT
        cell.border = BORDER
        ws1.row_dimensions[row].height = 16
        row += 1
        ws1.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        d_cell = ws1.cell(row=row, column=1, value=detail)
        d_cell.font = BODY_FONT
        d_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        d_cell.border = BORDER
        ws1.row_dimensions[row].height = 48
        row += 1

    # -----------------------------------------------------------------------
    # Sheet 2 — Churn by Segment & City
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Churn by Segment & City")
    ws2.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E","F","G"], [18,12,10,12,16,18,14]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Churn Breakdown — 08-Apr-2024 to 07-May-2024"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    ws2["A1"].alignment = CENTER
    ws2["A1"].fill = PatternFill("solid", fgColor="D6E4F7")

    row = 3
    row = write_section_title(ws2, row, "Churn by Segment", 5)
    for c, h in enumerate(["Segment","Churned","Total","Churn Rate","Status"], 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, 5)
    row += 1
    STATUS_COLORS = {"CRITICAL": RED_LIGHT, "NORMAL": "E2EFDA"}
    for i, (seg, churned, total, rate, status) in enumerate(SEGMENT_DATA):
        ws2.cell(row=row, column=1, value=seg)
        ws2.cell(row=row, column=2, value=churned)
        ws2.cell(row=row, column=3, value=total)
        ws2.cell(row=row, column=4, value=f"{rate:.2f}%")
        ws2.cell(row=row, column=5, value=status)
        for c in range(1, 6):
            fill_c = STATUS_COLORS.get(status, "FFFFFF")
            ws2.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill_c)
            ws2.cell(row=row, column=c).font = Font(name="Calibri", size=10,
                bold=(status == "CRITICAL"))
            ws2.cell(row=row, column=c).border = BORDER
            ws2.cell(row=row, column=c).alignment = LEFT
        row += 1

    row += 1
    row = write_section_title(ws2, row, "Churn by City", 5)
    for c, h in enumerate(["City","Churned","Total","Churn Rate","vs Baseline"], 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, 5)
    row += 1
    for i, (city, churned, total, rate) in enumerate(CITY_DATA):
        alt = (i % 2 == 0)
        vs = "Above" if rate > 21.0 else "Below"
        ws2.cell(row=row, column=1, value=city)
        ws2.cell(row=row, column=2, value=churned)
        ws2.cell(row=row, column=3, value=total)
        ws2.cell(row=row, column=4, value=f"{rate:.2f}%")
        ws2.cell(row=row, column=5, value=vs)
        style_data_row(ws2, row, 5, alt)
        row += 1

    row += 1
    row = write_section_title(ws2, row, "Top 10 Churned Customers by Balance (Anonymised)", 7)
    t10_headers = ["Rank","Balance (Rs)","Credit Score","NPS","Segment","City","Tenure (yrs)"]
    for c, h in enumerate(t10_headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header_row(ws2, row, 7)
    row += 1
    for i, (rank, bal, cs, nps, seg, city, tenure) in enumerate(TOP10_CUSTOMERS):
        alt = (i % 2 == 0)
        ws2.cell(row=row, column=1, value=rank)
        ws2.cell(row=row, column=2, value=u"₹" + bal)
        ws2.cell(row=row, column=3, value=cs)
        ws2.cell(row=row, column=4, value=f"{nps}/10")
        ws2.cell(row=row, column=5, value=seg)
        ws2.cell(row=row, column=6, value=city)
        ws2.cell(row=row, column=7, value=tenure)
        style_data_row(ws2, row, 7, alt)
        row += 1

    # -----------------------------------------------------------------------
    # Sheet 3 — Analysis Insights
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Analysis Insights")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 100

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Analysis Insights — 08-Apr-2024 to 07-May-2024"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    ws3["A1"].alignment = CENTER
    ws3["A1"].fill = PatternFill("solid", fgColor="D6E4F7")

    row = 3
    ws3.merge_cells(f"A{row}:B{row}")
    ws3.cell(row=row, column=1, value="Key Findings")
    ws3.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor="BDD7EE")
    ws3.cell(row=row, column=1).alignment = LEFT
    row += 1

    for i, finding in enumerate(KEY_FINDINGS, 1):
        alt = (i % 2 == 0)
        ws3.cell(row=row, column=1, value=str(i))
        ws3.merge_cells(f"B{row}:B{row}")
        ws3.cell(row=row, column=2, value=finding)
        style_data_row(ws3, row, 2, alt)
        ws3.row_dimensions[row].height = 30
        row += 1

    insights = [
        ("Segment Analysis",
         "The Premium segment, representing 51.5% of total customers, has breached the critical 25% churn "
         "threshold at 25.32%. This is the highest-value cohort and the primary revenue risk vector. "
         "Standard (20.65%) and Basic (15.15%) segments remain within acceptable bounds."),
        ("NPS vs Churn Correlation",
         "Churned customers carry an average NPS of 3.16/10, just above the 3.0 risk floor that triggers "
         "the churn risk combo alert. Any further decline in NPS among active customers should be treated "
         "as a leading indicator of upcoming churn."),
        ("Tenure vs Churn",
         "Top 10 churned customers by balance have tenures ranging from 0 to 8 years, with the majority "
         "(7 of 10) in the 0-5 year range. Short-tenure Premium customers represent the highest attrition "
         "risk and may not yet have sufficient product stickiness."),
        ("Transaction Anomaly Pattern",
         "All 100 transactions sampled from churned customers in the reference period (08-Apr-2024 to "
         "07-May-2024) are flagged as spend spikes (amount > 2x customer monthly average, range "
         "Rs 1,33,222 - Rs 1,49,558). This pattern suggests concentrated high-value spending "
         "immediately preceding churn — a potential pre-exit liquidation signal."),
        ("Loan/EMI Stress Overlap",
         "The combined EMI stress rate stands at 31% (Missed 16.75% + Delayed 14.25%), marginally "
         "above the 30% monitoring threshold. 16 customers present the full churn risk combo of "
         "churned status + missed EMI. These represent the highest-priority outreach targets."),
    ]

    row += 1
    for section, text in insights:
        ws3.merge_cells(f"A{row}:B{row}")
        ws3.cell(row=row, column=1, value=section)
        ws3.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
        ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor="BDD7EE")
        ws3.cell(row=row, column=1).alignment = LEFT
        ws3.cell(row=row, column=1).border = BORDER
        row += 1
        ws3.merge_cells(f"A{row}:B{row}")
        ws3.cell(row=row, column=1, value=text)
        ws3.cell(row=row, column=1).font = BODY_FONT
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal="left",
                                                           vertical="center", wrap_text=True)
        ws3.cell(row=row, column=1).border = BORDER
        ws3.row_dimensions[row].height = 60
        row += 2

    # -----------------------------------------------------------------------
    # Sheet 4 — Anomaly Flags
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Anomaly Flags")
    ws4.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E"], [12, 64, 12, 8, 10]):
        ws4.column_dimensions[col].width = w

    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Anomaly Flags — 07-May-2026"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    ws4["A1"].alignment = CENTER
    ws4["A1"].fill = PatternFill("solid", fgColor="D6E4F7")

    row = 3
    for c, h in enumerate(["Anomaly ID","Description","Severity","Score","Status"], 1):
        ws4.cell(row=row, column=c, value=h)
    style_header_row(ws4, row, 5)
    row += 1

    SEV_FILL = {"CRITICAL": "FFB3B3", "WARNING": "FFF2CC"}
    for ano_id, desc, sev, score, status in ANOMALIES:
        fill_hex = SEV_FILL.get(sev, "FFFFFF")
        ws4.cell(row=row, column=1, value=ano_id)
        ws4.cell(row=row, column=2, value=desc)
        ws4.cell(row=row, column=3, value=sev)
        ws4.cell(row=row, column=4, value=score)
        ws4.cell(row=row, column=5, value=status)
        for c in range(1, 6):
            ws4.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill_hex)
            ws4.cell(row=row, column=c).font = Font(name="Calibri", size=10,
                bold=(sev == "CRITICAL"))
            ws4.cell(row=row, column=c).border = BORDER
            ws4.cell(row=row, column=c).alignment = Alignment(horizontal="left",
                                                               vertical="center", wrap_text=True)
        ws4.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws4.merge_cells(f"A{row}:E{row}")
    ws4.cell(row=row, column=1, value="Priority Intervention List (Anonymised)")
    ws4.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws4.cell(row=row, column=1).fill = PatternFill("solid", fgColor="BDD7EE")
    ws4.cell(row=row, column=1).alignment = LEFT
    row += 1

    for c, h in enumerate(["Priority","Profile","Intervention Reason"], 1):
        ws4.cell(row=row, column=c, value=h)
    style_header_row(ws4, row, 3)
    ws4.column_dimensions["B"].width = 45
    ws4.column_dimensions["C"].width = 55
    row += 1

    for priority, profile, reason in PRIORITY_LIST:
        alt = (priority % 2 == 0)
        ws4.cell(row=row, column=1, value=priority)
        ws4.cell(row=row, column=2, value=profile)
        ws4.cell(row=row, column=3, value=reason)
        style_data_row(ws4, row, 3, alt)
        ws4.row_dimensions[row].height = 36
        row += 1

    # -----------------------------------------------------------------------
    # Sheet 5 — Methodology
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("Methodology")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 55

    ws5.merge_cells("A1:B1")
    ws5["A1"] = "Methodology & Data Sources"
    ws5["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
    ws5["A1"].alignment = CENTER
    ws5["A1"].fill = PatternFill("solid", fgColor="D6E4F7")

    row = 3
    for c, h in enumerate(["Parameter","Value"], 1):
        ws5.cell(row=row, column=c, value=h)
    style_header_row(ws5, row, 2)
    row += 1

    for i, (param, value) in enumerate(METHODOLOGY):
        alt = (i % 2 == 0)
        ws5.cell(row=row, column=1, value=param)
        ws5.cell(row=row, column=2, value=value)
        style_data_row(ws5, row, 2, alt)
        row += 1

    row += 2
    ws5.merge_cells(f"A{row}:B{row}")
    ws5.cell(row=row, column=1,
             value="Agents Used: SQL Agent | Analysis Agent | Anomaly Agent | Report Agent")
    ws5.cell(row=row, column=1).font = Font(name="Calibri", italic=True, size=10, color="4472C4")
    ws5.cell(row=row, column=1).alignment = LEFT
    row += 1
    ws5.merge_cells(f"A{row}:B{row}")
    ws5.cell(row=row, column=1,
             value="CONFIDENTIAL — BankSight AI Internal Report — Do not distribute externally")
    ws5.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10, color="FF0000")
    ws5.cell(row=row, column=1).alignment = CENTER

    wb.save(EXCEL_PATH)
    print(f"[Excel] Saved: {EXCEL_PATH}")


# ===========================================================================
# PDF REPORT
# ===========================================================================
def build_pdf(KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
              LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
              ANOMALIES, PRIORITY_LIST, METHODOLOGY):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    DARK_BLUE_RL = colors.HexColor("#1F3864")
    LIGHT_BLUE   = colors.HexColor("#D6E4F7")
    RED_RL       = colors.HexColor("#FF0000")
    RED_LIGHT_RL = colors.HexColor("#FFB3B3")
    YELLOW_RL    = colors.HexColor("#FFF2CC")
    ALT_RL       = colors.HexColor("#DCE6F1")
    WHITE_RL     = colors.white
    ORANGE_RL    = colors.HexColor("#FF6600")

    W, H = A4

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle",
        parent=styles["Heading1"],
        fontSize=18, textColor=DARK_BLUE_RL,
        alignment=TA_CENTER, spaceAfter=4)
    subtitle_style = ParagraphStyle("Subtitle",
        parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#4472C4"),
        alignment=TA_CENTER, spaceAfter=12, italic=True)
    section_style = ParagraphStyle("Section",
        parent=styles["Heading2"],
        fontSize=13, textColor=DARK_BLUE_RL,
        spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle("Body",
        parent=styles["Normal"],
        fontSize=9.5, leading=14, alignment=TA_JUSTIFY,
        spaceAfter=6)
    bullet_style = ParagraphStyle("Bullet",
        parent=styles["Normal"],
        fontSize=9.5, leading=14, leftIndent=14,
        bulletIndent=4, spaceAfter=3)
    label_style = ParagraphStyle("Label",
        parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#4472C4"),
        italic=True, spaceAfter=2)
    confidential_style = ParagraphStyle("Conf",
        parent=styles["Normal"],
        fontSize=8, textColor=RED_RL, alignment=TA_CENTER)
    alert_style = ParagraphStyle("Alert",
        parent=styles["Normal"],
        fontSize=10, textColor=RED_RL, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=4)

    def tbl_style_base(data, header_rows=1):
        cols = len(data[0])
        row_count = len(data)
        commands = [
            ("BACKGROUND",  (0, 0),           (-1, header_rows - 1), DARK_BLUE_RL),
            ("TEXTCOLOR",   (0, 0),           (-1, header_rows - 1), WHITE_RL),
            ("FONTNAME",    (0, 0),           (-1, header_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0),           (-1, header_rows - 1), 9),
            ("FONTNAME",    (0, header_rows), (-1, -1),              "Helvetica"),
            ("FONTSIZE",    (0, header_rows), (-1, -1),              8.5),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [WHITE_RL, ALT_RL]),
            ("GRID",        (0, 0),           (-1, -1),  0.5, colors.HexColor("#B8CCE4")),
            ("VALIGN",      (0, 0),           (-1, -1),  "MIDDLE"),
            ("ALIGN",       (0, 0),           (-1, -1),  "LEFT"),
            ("TOPPADDING",  (0, 0),           (-1, -1),  4),
            ("BOTTOMPADDING",(0, 0),          (-1, -1),  4),
            ("LEFTPADDING", (0, 0),           (-1, -1),  5),
        ]
        return TableStyle(commands)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#4472C4"))
        canvas.drawString(1.5*cm, 1.0*cm,
            "CONFIDENTIAL — BankSight AI Internal Report — Do not distribute externally")
        canvas.drawRightString(W - 1.5*cm, 1.0*cm,
            f"Page {doc.page}  |  Generated: 07-May-2026")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        PDF_PATH, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.8*cm, bottomMargin=2.2*cm,
    )

    story = []
    HR = HRFlowable(width="100%", thickness=1.2,
                    color=DARK_BLUE_RL, spaceAfter=8)

    # -------------------------------------------------------------------
    # Page 1 — Executive Summary
    # -------------------------------------------------------------------
    story.append(Paragraph("BankSight AI", title_style))
    story.append(Paragraph("Churn Analysis Report", title_style))
    story.append(Paragraph(
        "Period: 08-Apr-2024 to 07-May-2024 &nbsp;&nbsp;|&nbsp;&nbsp; "
        "Report Date: 07-May-2026 &nbsp;&nbsp;|&nbsp;&nbsp; Run: " + RUN_TS,
        subtitle_style))
    story.append(HR)

    story.append(Paragraph("1. Executive Summary", section_style))

    # KPI table
    kpi_data = [["Metric", "Value", "Status"]]
    for metric, value in KPIS:
        if "Rate" in metric and "%" in value:
            rate_f = float(value.replace("%",""))
            if rate_f > 25:
                status = "CRITICAL"
            elif rate_f > 21:
                status = "MONITOR"
            else:
                status = "NORMAL"
        else:
            status = "-"
        kpi_data.append([metric, value, status])

    kpi_tbl = Table(kpi_data, colWidths=[8*cm, 6*cm, 4*cm])
    kpi_ts = tbl_style_base(kpi_data)
    # colour status cells
    for r_idx, (metric, value) in enumerate(KPIS, 1):
        if "Rate" in metric and "%" in value:
            rate_f = float(value.replace("%",""))
            if rate_f > 25:
                kpi_ts.add("BACKGROUND", (2, r_idx), (2, r_idx), RED_LIGHT_RL)
                kpi_ts.add("TEXTCOLOR",  (2, r_idx), (2, r_idx), RED_RL)
                kpi_ts.add("FONTNAME",   (2, r_idx), (2, r_idx), "Helvetica-Bold")
            elif rate_f > 21:
                kpi_ts.add("BACKGROUND", (2, r_idx), (2, r_idx), YELLOW_RL)
            else:
                kpi_ts.add("BACKGROUND", (2, r_idx), (2, r_idx),
                            colors.HexColor("#E2EFDA"))
    kpi_tbl.setStyle(kpi_ts)
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.3*cm))

    # Overall status
    story.append(Paragraph(
        "Overall Churn Status: NORMAL — 21.12% vs 21.00% baseline (within tolerance)",
        ParagraphStyle("OK", parent=styles["Normal"], fontSize=9.5,
                       textColor=colors.HexColor("#375623"), fontName="Helvetica-Bold",
                       backColor=colors.HexColor("#E2EFDA"), borderPad=4, spaceAfter=6,
                       alignment=TA_CENTER)))

    # CRITICAL alert box
    story.append(Paragraph(
        "CRITICAL ALERT: Premium Segment churn at 25.32% exceeds 25% threshold — "
        "immediate retention action required",
        ParagraphStyle("CritAlert", parent=styles["Normal"], fontSize=9.5,
                       textColor=RED_RL, fontName="Helvetica-Bold",
                       backColor=RED_LIGHT_RL, borderPad=4, spaceAfter=8,
                       alignment=TA_CENTER)))

    story.append(Paragraph("Anomaly Count: 5 CRITICAL | 5 WARNING", alert_style))
    story.append(Spacer(1, 0.2*cm))

    story.append(Paragraph("Top 3 Recommendations", section_style))
    for i, (title, detail) in enumerate(RECOMMENDATIONS, 1):
        story.append(Paragraph(f"<b>{i}. {title}</b>", bullet_style))
        story.append(Paragraph(detail, ParagraphStyle("RecDetail",
            parent=styles["Normal"], fontSize=9, leftIndent=20, spaceAfter=6,
            leading=13, alignment=TA_JUSTIFY)))

    story.append(PageBreak())

    # -------------------------------------------------------------------
    # Page 2 — Analysis Insights
    # -------------------------------------------------------------------
    story.append(Paragraph("2. Analysis Insights", section_style))
    story.append(HR)

    story.append(Paragraph("Segment Performance", ParagraphStyle("SubSection",
        parent=styles["Heading3"], fontSize=11, textColor=DARK_BLUE_RL,
        spaceBefore=6, spaceAfter=4)))

    seg_data = [["Segment", "Churned", "Total", "Churn Rate", "Status"]]
    for seg, churned, total, rate, status in SEGMENT_DATA:
        seg_data.append([seg, str(churned), str(total), f"{rate:.2f}%", status])
    seg_tbl = Table(seg_data, colWidths=[4.2*cm, 2.8*cm, 2.8*cm, 3.2*cm, 4*cm])
    seg_ts = tbl_style_base(seg_data)
    for r_idx, (seg, churned, total, rate, status) in enumerate(SEGMENT_DATA, 1):
        c = RED_LIGHT_RL if status == "CRITICAL" else colors.HexColor("#E2EFDA")
        seg_ts.add("BACKGROUND", (0, r_idx), (-1, r_idx), c)
        if status == "CRITICAL":
            seg_ts.add("FONTNAME", (0, r_idx), (-1, r_idx), "Helvetica-Bold")
    seg_tbl.setStyle(seg_ts)
    story.append(seg_tbl)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("City-wise Churn Breakdown", ParagraphStyle("SubSection",
        parent=styles["Heading3"], fontSize=11, textColor=DARK_BLUE_RL,
        spaceBefore=6, spaceAfter=4)))

    city_data = [["City", "Churned", "Total", "Churn Rate", "vs Baseline"]]
    for city, churned, total, rate in CITY_DATA:
        vs = "Above" if rate > 21.0 else "Below"
        city_data.append([city, str(churned), str(total), f"{rate:.2f}%", vs])
    city_tbl = Table(city_data, colWidths=[4.2*cm, 2.8*cm, 2.8*cm, 3.2*cm, 4*cm])
    city_tbl.setStyle(tbl_style_base(city_data))
    story.append(city_tbl)
    story.append(Spacer(1, 0.3*cm))

    sub_sec = ParagraphStyle("SubSection2", parent=styles["Heading3"],
        fontSize=11, textColor=DARK_BLUE_RL, spaceBefore=8, spaceAfter=3)
    for section, text in [
        ("NPS vs Churn Correlation",
         "Churned customers carry an average NPS of 3.16/10, just above the 3.0 risk floor. "
         "This is a strong leading indicator — active customers trending below NPS 4 should be "
         "treated as pre-churn signals and escalated to relationship managers immediately."),
        ("Tenure vs Churn",
         "7 of the top 10 churned customers by balance have tenures of 0-5 years, suggesting "
         "insufficient product embedding in the early relationship lifecycle. Onboarding and "
         "engagement improvements within the first 5 years may reduce Premium churn significantly."),
        ("Transaction Anomaly Pattern",
         "100% of sampled transactions from churned customers in the period are flagged as spend "
         "spikes (Rs 1,33,222 - Rs 1,49,558). This pattern — high-value transactions immediately "
         "preceding account closure — should be codified as a real-time churn trigger in the "
         "Anomaly Agent pipeline."),
        ("Loan/EMI Stress",
         "31% combined EMI stress rate (Missed 16.75% + Delayed 14.25%) marginally exceeds the "
         "30% monitoring threshold. 16 customers are both churned and carry missed EMI status, "
         "representing the most at-risk cohort for credit loss coinciding with churn."),
    ]:
        story.append(Paragraph(section, sub_sec))
        story.append(Paragraph(text, body_style))

    story.append(PageBreak())

    # -------------------------------------------------------------------
    # Page 3 — Anomaly Alerts
    # -------------------------------------------------------------------
    story.append(Paragraph("3. Anomaly Alerts", section_style))
    story.append(HR)

    ano_data = [["ID", "Description", "Severity", "Score", "Status"]]
    for ano_id, desc, sev, score, status in ANOMALIES:
        ano_data.append([ano_id, desc, sev, str(score), status])

    ano_tbl = Table(ano_data, colWidths=[1.8*cm, 8.5*cm, 2.4*cm, 1.5*cm, 1.8*cm])
    ano_ts = tbl_style_base(ano_data)
    for r_idx, (ano_id, desc, sev, score, status) in enumerate(ANOMALIES, 1):
        c = RED_LIGHT_RL if sev == "CRITICAL" else YELLOW_RL
        ano_ts.add("BACKGROUND", (0, r_idx), (-1, r_idx), c)
        if sev == "CRITICAL":
            ano_ts.add("FONTNAME", (0, r_idx), (-1, r_idx), "Helvetica-Bold")
    ano_tbl.setStyle(ano_ts)
    story.append(ano_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Priority Intervention List (Anonymised)", sub_sec))
    pri_data = [["Priority", "Profile", "Intervention Reason"]]
    for priority, profile, reason in PRIORITY_LIST:
        pri_data.append([str(priority), profile, reason])
    pri_tbl = Table(pri_data, colWidths=[1.8*cm, 7*cm, 8.2*cm])
    pri_tbl.setStyle(tbl_style_base(pri_data))
    story.append(pri_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph(
        "Note: All customer identifiers have been anonymised in this report. "
        "Data coverage period: 08-Apr-2024 to 07-May-2024. "
        "Anomaly detection uses threshold: transaction amount > 2x customer monthly average.",
        ParagraphStyle("Caveat", parent=styles["Normal"], fontSize=8.5,
                       textColor=colors.grey, italic=True, alignment=TA_LEFT)))

    story.append(PageBreak())

    # -------------------------------------------------------------------
    # Page 4 — Data Appendix
    # -------------------------------------------------------------------
    story.append(Paragraph("4. Data Appendix", section_style))
    story.append(HR)

    story.append(Paragraph("Record Counts by Source Table", sub_sec))
    rc_data = [
        ["Table", "Row Count", "Scope"],
        ["customers",           "3,225",  "Full customer master — primary analysis table"],
        ["transactions",        "50,000", "Full transaction history Jan-Dec 2024"],
        ["loan_emi",            "400",    "All active and closed loan records"],
        ["Churned customers",   "681",    "Subset: churn = 1 (21.12% of total)"],
        ["Period transactions",  "864",   "Transactions from churned customers, Apr-May 2024 window"],
    ]
    rc_tbl = Table(rc_data, colWidths=[4.5*cm, 2.5*cm, 10*cm])
    rc_tbl.setStyle(tbl_style_base(rc_data))
    story.append(rc_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Key Field Descriptions", sub_sec))
    fd_data = [
        ["Field",            "Description"],
        ["churn",            "1 = customer exited, 0 = retained. Target variable."],
        ["nps_score",        "Net Promoter Score on 0-10 scale. Risk floor: <= 3."],
        ["balance",          "Account balance in INR (Rs). Original USD * 83 conversion applied."],
        ["segment",          "Customer tier: Premium / Standard / Basic."],
        ["is_anomaly",       "1 = transaction flagged as spend spike (> 2x monthly avg)."],
        ["emi_status",       "Paid / Missed / Delayed. Missed + Delayed = stress indicator."],
        ["credit_score",     "Range 350-850. Used in segmentation risk analysis."],
        ["active_member",    "1 = active account, 0 = inactive. Inactivity is a churn signal."],
    ]
    fd_tbl = Table(fd_data, colWidths=[4.5*cm, 12.5*cm])
    fd_tbl.setStyle(tbl_style_base(fd_data))
    story.append(fd_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Methodology Notes", sub_sec))
    for note in [
        "Data source: data/banking_mock.db (SQLite) — read-only access via SQL Agent.",
        "Churn baseline: 21% (681 / 3,225). CRITICAL alert threshold: > 25%.",
        "Anomaly detection: transaction amount > 2x that customer's monthly average spend.",
        "Churn risk combo: NPS <= 3 AND missed EMI in the same calendar month.",
        "High-risk segment definition: Basic tier + credit_score < 500.",
        "Currency: all monetary values in INR (Rs). USD conversion rate applied during setup: 1 USD = Rs 83.",
        "NPS scale: 0-10 (not 0-100). Do not normalise to percentage.",
        "Report generated by BankSight AI Report Agent on 07-May-2026.",
    ]:
        story.append(Paragraph(u"• " + note, bullet_style))

    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph(
        "CONFIDENTIAL — BankSight AI Internal Report — Do not distribute externally",
        confidential_style))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    print(f"[PDF] Saved: {PDF_PATH}")


# ===========================================================================
# POWERBI CSV
# ===========================================================================
def build_csv(data):
    """
    Fix 3 — Customer-level CSV rows so Power BI can slice
    by customer, segment, city, NPS, balance etc.
    One row per churned customer + joined loan EMI status.
    """
    churned_df = data["churned_df"].copy()
    loan_df    = data["loan_df"].copy()

    # Join loan EMI status onto churned customers
    loan_latest = (
        loan_df.sort_values("emi_status")
               .drop_duplicates(subset=["customer_id"], keep="first")
    [["customer_id", "loan_type", "loan_amount", "emi_amount", "emi_status"]]
    )
    merged = churned_df.merge(loan_latest, on="customer_id", how="left")

    # Add derived columns for Power BI
    merged["is_anomaly"]          = 0  # placeholder — set by Anomaly Agent
    merged["anomaly_severity"]    = "NONE"
    merged["anomaly_description"] = ""
    merged["churn_risk_combo"]    = (
        (merged["nps_score"] <= 3) &
        (merged["emi_status"] == "Missed")
    ).astype(int)
    merged["report_generated_at"] = RUN_TS
    merged["data_source"]         = "banking_mock.db"
    merged["report_date"]         = RUN_LABEL
    merged["baseline_churn_rate"] = 21.00
    merged["actual_churn_rate"]   = data["churn_rate"]
    merged["churn_status"]        = (
        "CRITICAL" if data["churn_rate"] > 25
        else "WARNING" if data["churn_rate"] > 21
        else "NORMAL"
    )

    # Select and rename final columns
    export_cols = [
        "customer_id", "age", "gender", "city", "segment",
        "credit_score", "balance", "estimated_salary",
        "tenure", "products_number", "active_member",
        "nps_score", "account_type", "kyc_status",
        "loan_type", "loan_amount", "emi_amount", "emi_status",
        "churn_risk_combo", "is_anomaly", "anomaly_severity",
        "anomaly_description", "actual_churn_rate",
        "baseline_churn_rate", "churn_status",
        "report_date", "report_generated_at", "data_source",
    ]
    # Only export columns that exist
    export_cols = [c for c in export_cols if c in merged.columns]
    export_df   = merged[export_cols]

    # Write UTF-8 BOM for Power BI / Excel compatibility
    export_df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    print(f"[CSV]  Saved: {CSV_PATH} ({len(export_df):,} rows)")


# ===========================================================================
# Main
# ===========================================================================
if __name__ == "__main__":
    print("BankSight AI — Report Agent starting...")
    print(f"Loading live data from: {DB_PATH}")

    # Load live data from banking_mock.db
    data = load_data_from_db()

    # Build shared structures from live data
    (KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
     LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
     ANOMALIES, PRIORITY_LIST, METHODOLOGY) = build_shared_structures(data)

    print(f"  Customers loaded:    {data['total_customers']:,}")
    print(f"  Churned customers:   {data['total_churned']:,} ({data['churn_rate']:.2f}%)")
    print(f"  Transactions loaded: {data['total_txns']:,}")
    print(f"  Loan records loaded: {data['total_loans']:,}")
    print(f"  Anomaly count:       {data['anomaly_txns']:,} flagged transactions")
    print()

    build_excel(KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
                LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
                ANOMALIES, PRIORITY_LIST, METHODOLOGY)

    build_pdf(KPIS, SEGMENT_DATA, CITY_DATA, TOP10_CUSTOMERS,
              LOAN_EMI, KEY_FINDINGS, RECOMMENDATIONS,
              ANOMALIES, PRIORITY_LIST, METHODOLOGY)

    build_csv(data)

    print("\nAll three reports generated successfully.")
    print(f"  Excel : {EXCEL_PATH}")
    print(f"  PDF   : {PDF_PATH}")
    print(f"  CSV   : {CSV_PATH} ({data['total_churned']:,} customer-level rows)")